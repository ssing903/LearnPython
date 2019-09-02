# Chapter12.py Working with Excel Spreadsheets

# Installing the openpyxl modules

'''
import openpyxl,os
print(os.getcwd())

wb = openpyxl.load_workbook('AutomateBoringStuffWithPython/Book1.xlsx')
print(wb.get_active_sheet().title)
'''

# Getting Cells from the sheets
# import openpyxl

# wb = openpyxl.load_workbook('AutomateBoringStuffWithPython/Book1.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet['A1'].value)

# A3 = sheet['A3']
# print(A3.value)

# A3 = 'Row ' + str(A3.row) + ', column ' + str(A3.coordinate) + ' is ' + str(A3.value)
# print(A3)
# print(sheet.cell(row=1,column=2).value)

# for i in range(1,4,1):
#      print(sheet.cell(row=i+1,column=2).value)

# Converting between columns and letters

# from openpyxl.utils import get_column_letter, column_index_from_string

# print(get_column_letter(1))
'''
print(wb.worksheets[0].max_row)
print(wb.worksheets[0].max_column)
'''

'''
# Getting row and column from the sheets
# print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
     # print(rowOfCellObjects)
     for x in rowOfCellObjects:
          print(x.value,sep='\n',end=' ')
     print(' ')

     # for cellObj in rowOfCellObjects:
     #      print(cellObj.coordinate,cellObj.value)
     # print('End of row')
'''

# print(len(wb.worksheets))

# Workbooks sheets and cells,
# Reading data from the spreadsheet

# Step I read the spreadsheet data
'''
import openpyxl, pprint

print('Opening workbook')
wb = openpyxl.load_workbook('AutomateBoringStuffWithPython/censuspopdata.xlsx')
sheet = wb.worksheets[0]
countyData  = {}

# Fill in county data with each county's population and tracts
print('Reading rows. ')

# for row in range(2,4):
for row in range(2,sheet.max_row + 1):
     # Each row in the spreadsheet has data for one census tract
     state = sheet['B' + str(row)].value
     county = sheet['C' + str(row)].value
     pop = sheet['D' + str(row)].value

     # make sure that key for state exists
     countyData.setdefault(state,{})
     countyData[state].setdefault(county,{'tracts':0,'pop':0})

     # Each row represents one census tract, so increment by one
     countyData[state][county]['tracts'] += 1
     countyData[state][county]['pop'] += int(pop)

# Open a text document and write a data to it
print('Writing Results')
resultFile  = open('census2010.py','w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
'''
# import census2010
# import census2010

# print(os.getcwd())
# print(census2010.allData['AK']['Anchorage'])

# Writing excel documents
'''
import openpyxl

wb = openpyxl.Workbook()
wb.worksheets[0].title = 'Hello World'
print(wb.worksheets[0].title)
wb.save('HelloWorld.xlsx')
'''

# Project updating spreedsheets

